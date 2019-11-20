
import openpyxl, smtplib, sys
# Open the spreadsheet and get the latest dues status.

wb = openpyxl.load_workbook('UserReminders.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')


lastCol= sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value


dueUsers = {}
# Check each member's due status
for r in range(2, sheet.max_row + 1):
   oil = sheet.cell(row=r, column=lastCol).value
   if oil != 'Due':
       name = sheet.cell(row=r, column=1).value
       email = sheet.cell(row=r, column=2).value
       dueUsers[name] = email

# Log in to email account.
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('my_email_address@gmail.com', sys.argv [1])


# Send out reminder emails.
for name, email in dueUsers.items():
   body = 'Subject: %s User Notification Reminder.\nDear %s,\nRecords show that you have reached a certain mileage and your oil need to be changed %s. Please make an appointment with us as soon as Possible. Thank you!' % (latestMonth, name, latestMonth)
   print('Sending email to %s...' % email)
   sendmailStatus = smtpObj.sendmail('my_email_address@gmail.com', email, body)

   if sendmailStatus != {}:
       print('There was a problem sending email to %s: %s' % (email, sendmailStatus))

smtpObj.quit()
